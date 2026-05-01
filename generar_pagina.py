from openpyxl import load_workbook
from urllib.parse import quote
from pathlib import Path
import unicodedata
import re

WHATSAPP_NUMERO = "5491163509142"  # Celular WhatsApp sin + ni espacios.
ARCHIVO_PRODUCTOS = "productos.xlsx"

MARCAS_ORDEN = ["Ana Grant", "Aretha", "BLUO", "Deville", "So Pink", "Stylo"]

GRUPOS_CATEGORIAS = {
    "Lencería": ["Ropa de dormir", "Medias"],
    "Corsetería": [
        "Corpiños",
        "Bombachas",
        "Portaligas y ligas",
        "Corset",
        "Body",
        "Bra importado",
    ],
}


def normalizar(texto):
    texto = str(texto or "").strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.lower()


def contiene(texto, palabras):
    texto_normalizado = normalizar(texto)
    return any(palabra in texto_normalizado for palabra in palabras)


def marca_carpeta(marca):
    mapa = {
        "ana grant": "Ana Grant",
        "aretha": "Aretha",
        "bluo": "BLUO",
        "deville": "Deville",
        "so pink": "So Pink",
        "so pink!": "So Pink",
        "stylo": "Stylo",
    }
    return mapa.get(normalizar(marca), str(marca).strip())


def es_destacado(valor):
    return normalizar(valor) in ["si", "sí", "s", "x", "true", "1", "destacado"]


def obtener_descuento(valor):
    """
    Lee la columna Oferta del Excel.
    Ejemplos válidos: "oferta 10%", "oferta 15", "20%" o "20".
    """
    texto = normalizar(valor)
    if not texto:
        return 0

    match = re.search(r"(\d{1,2})", texto)
    if not match:
        return 0

    porcentaje = int(match.group(1))
    if porcentaje <= 0 or porcentaje >= 100:
        return 0

    return porcentaje


def precio_a_numero(precio):
    texto = str(precio or "").strip()
    if not texto:
        return None

    texto = texto.replace("$", "").replace(" ", "")

    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(".", "")

    try:
        return float(texto)
    except ValueError:
        return None


def formatear_pesos(valor):
    if valor is None:
        return ""

    entero = int(round(valor))
    return "$" + f"{entero:,}".replace(",", ".")


def calcular_precio_oferta(precio, descuento):
    numero = precio_a_numero(precio)
    if numero is None or descuento <= 0:
        return ""

    return formatear_pesos(numero * (1 - descuento / 100))


def precio_html(producto):
    descuento = producto.get("descuento", 0)
    precio = producto.get("precio", "")

    if descuento > 0:
        precio_oferta = producto.get("precio_oferta") or calcular_precio_oferta(precio, descuento)
        return f"""
          <p class="precio precio-con-oferta">
            <span class="precio-anterior">{precio}</span>
            <span class="precio-oferta">{precio_oferta}</span>
            <span class="descuento-pill">−{descuento}%</span>
          </p>
        """

    return f'<p class="precio">{precio}</p>'


def producto_visible_por_stock(valor):
    """
    Controla desde Excel si un producto se muestra o no.

    En la columna Stock podés poner:
    - "No", "Sin stock", "Agotado", "No disponible" o "0" para ocultarlo.
    - "Sí", "Disponible", un número mayor a 0 o dejarlo vacío para mostrarlo.

    También oculta variantes como "SIN STOCK", "sin-stock",
    "agotado momentáneamente" o "no disponible".
    """
    if valor is None:
        return True

    if isinstance(valor, (int, float)):
        return valor > 0

    stock = normalizar(valor)
    stock_simple = re.sub(r"[^a-z0-9]+", " ", stock).strip()

    if stock_simple == "":
        return True

    ocultar_exactos = {
        "0",
        "no",
        "n",
        "false",
        "falso",
        "sin stock",
        "stock 0",
        "no stock",
        "no disponible",
        "no hay stock",
    }

    ocultar_si_contiene = [
        "sin stock",
        "agotado",
        "agotada",
        "agotados",
        "agotadas",
        "no disponible",
        "discontinuado",
        "discontinuada",
    ]

    if stock_simple in ocultar_exactos:
        return False

    if any(texto in stock_simple for texto in ocultar_si_contiene):
        return False

    return True


def clasificar_producto(producto):
    """
    Clasifica cada producto usando las columnas B, C y D del Excel:
    Nombre, Descripción y Rubro.
    """
    texto = " ".join([
        producto.get("nombre", ""),
        producto.get("descripcion", ""),
        producto.get("rubro", ""),
    ])

    reglas = [
        ("Lencería", "Ropa de dormir", [
            "camison", "bata", "pijama", "pijamas", "pantufla", "pantuflas",
            "pantuflon", "remeron", "musculosa", "remera", "short",
            "camiseta", "bermuda"
        ]),
        ("Lencería", "Medias", [
            "casual", "soquete", "canoa", "media", "medias", "invisible",
            "manguita"
        ]),
        ("Corsetería", "Portaligas y ligas", [
            "portaliga", "portaligas", "liga", "ligas"
        ]),
        ("Corsetería", "Corset", [
            "corset", "faja"
        ]),
        ("Corsetería", "Body", [
            "body", "bodies"
        ]),
        ("Corsetería", "Bra importado", [
            "tasa de silicona", "tasas de silicona", "taza de silicona",
            "tazas de silicona", "tasa de siliconas", "tasas de siliconas",
            "taza de siliconas", "tazas de siliconas", "pesonera",
            "pesoneras", "boop tape", "boob tape", "body tape"
        ]),
        ("Corsetería", "Bombachas", [
            "bikini", "colaless", "cola les", "culotte", "culote",
            "culotteless", "bombacha", "bombachas", "trusa", "brief",
            "vedetina", "tiro corto", "tanga"
        ]),
        ("Corsetería", "Corpiños", [
            "soutien", "soutiens", "corpiño", "corpino", "corpiños",
            "corpinos", "taza", "tazas", "triangulo", "triángulo",
            "bandeau"
        ]),
    ]

    for grupo, categoria, palabras in reglas:
        if contiene(texto, palabras):
            return grupo, categoria

    return "Otros", "Otros"


def buscar_imagen(codigo, marca):
    carpeta = Path("imagenes") / marca_carpeta(marca)
    if not carpeta.exists():
        return "imagenes/logo-mucho-de-ti.jpg"

    codigo_limpio = str(codigo).strip().lower()
    extensiones = ["*.jpg", "*.jpeg", "*.png", "*.webp"]

    for patron in extensiones:
        for archivo in carpeta.glob(patron):
            nombre = archivo.name.lower().strip()
            # Vincula archivos como: "109. Ana Grant.jpg", "109 Ana Grant.jpg", "109-Ana Grant.jpg"
            if re.match(rf"^{re.escape(codigo_limpio)}(\D|$)", nombre):
                return archivo.as_posix()

    return "imagenes/logo-mucho-de-ti.jpg"


def leer_productos_excel():
    wb = load_workbook(ARCHIVO_PRODUCTOS, data_only=True)
    ws = wb.active
    encabezados = [str(c.value or "").strip().lower() for c in ws[1]]

    def col(nombre, requerido=True):
        nombre = nombre.lower()
        if nombre in encabezados:
            return encabezados.index(nombre)
        if requerido:
            raise ValueError(f"Falta la columna obligatoria '{nombre}' en {ARCHIVO_PRODUCTOS}")
        return None

    productos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not any(fila):
            continue

        producto = {
            "codigo": str(fila[col("codigo")] or "").strip(),
            "nombre": str(fila[col("nombre")] or "").strip(),
            "descripcion": str(fila[col("descripcion")] or "").strip(),
            "rubro": str(fila[col("rubro")] or "").strip(),
            "marca": marca_carpeta(fila[col("marca")]),
            "precio": str(fila[col("precio")] or "").strip(),
            "talles": str(fila[col("talles")] or "Consultar disponibilidad").strip(),
            "stock": str(fila[col("stock")] or "").strip(),
            "destacado": es_destacado(fila[col("destacado")]),
        }

        oferta_col = col("oferta", requerido=False)
        oferta_valor = fila[oferta_col] if oferta_col is not None else ""
        descuento = obtener_descuento(oferta_valor)

        producto["oferta"] = str(oferta_valor or "").strip()
        producto["descuento"] = descuento
        producto["precio_oferta"] = calcular_precio_oferta(producto["precio"], descuento)

        if producto["codigo"] and producto["nombre"]:
            grupo, categoria = clasificar_producto(producto)
            producto["grupo"] = grupo
            producto["categoria"] = categoria
            producto["imagen"] = buscar_imagen(producto["codigo"], producto["marca"])
            productos.append(producto)

    return productos


# Leemos todos los productos del Excel y ocultamos los que estén marcados sin stock.
productos = [p for p in leer_productos_excel() if producto_visible_por_stock(p["stock"])]
productos_destacados = [p for p in productos if p["destacado"]]

nav_marcas = ""
for marca in MARCAS_ORDEN:
    nav_marcas += f'<a href="#productos" class="filtro-marca" data-marca="{marca}">{marca}</a>'

nav_categorias = ""
for grupo, categorias in GRUPOS_CATEGORIAS.items():
    botones = ""
    for categoria in categorias:
        botones += f'<a href="#productos" class="filtro-categoria" data-grupo="{grupo}" data-categoria="{categoria}">{categoria}</a>'

    nav_categorias += f"""
      <div class="grupo-categorias">
        <h3>{grupo}</h3>
        <div class="categoria-botones">
          {botones}
        </div>
      </div>
    """

nav_dropdown = '<a href="#productos" class="dropdown-destacados">Destacados</a>'
for grupo, categorias in GRUPOS_CATEGORIAS.items():
    nav_dropdown += f'<span>{grupo}</span>'
    for categoria in categorias:
        nav_dropdown += f'<a href="#productos" class="filtro-categoria" data-grupo="{grupo}" data-categoria="{categoria}">{categoria}</a>'

cards = ""
cards_ofertas = ""

for p in productos:
    mensaje = quote(f"Hola, quiero consultar por {p['nombre']} - Código {p['codigo']}")
    link = f"https://wa.me/{WHATSAPP_NUMERO}?text={mensaje}"
    destacado = "si" if p["destacado"] else "no"
    precio_producto = precio_html(p)
    clase_oferta = " oferta-card" if p["descuento"] > 0 else ""
    badge_oferta = '<span class="badge-oferta">OFERTA</span>' if p["descuento"] > 0 else ""
    precio_final = p["precio_oferta"] if p["descuento"] > 0 and p["precio_oferta"] else p["precio"]

    # Datos seguros para el carrito.
    data_codigo = str(p["codigo"]).replace('"', "&quot;")
    data_nombre = str(p["nombre"]).replace('"', "&quot;")
    data_marca = str(p["marca"]).replace('"', "&quot;")
    data_precio = str(precio_final).replace('"', "&quot;")
    data_imagen = str(p["imagen"]).replace('"', "&quot;")

    card_html = f"""
      <article class="card{clase_oferta}" data-marca="{p['marca']}" data-destacado="{destacado}" data-grupo="{p['grupo']}" data-categoria="{p['categoria']}" data-stock="si" data-oferta="{'si' if p['descuento'] > 0 else 'no'}">
        <div class="card-img-wrap">
          {badge_oferta}
          <img class="producto-img" src="{p['imagen']}" alt="{p['nombre']}" loading="lazy" onerror="this.src='imagenes/logo-mucho-de-ti.jpg'">
        </div>
        <div class="card-body">
          <span class="codigo">{p['marca']} · Cód. {p['codigo']}</span>
          <span class="etiqueta-categoria">{p['grupo']} · {p['categoria']}</span>
          <h3>{p['nombre']}</h3>
          <p class="descripcion">{p['descripcion']}</p>
          <p class="talles">{p['talles']}</p>
          {precio_producto}
          <button class="btn-carrito-card" type="button" data-codigo="{data_codigo}" data-nombre="{data_nombre}" data-marca="{data_marca}" data-precio="{data_precio}" data-imagen="{data_imagen}">
            AGREGAR AL CARRITO
          </button>
          <a class="consultar" href="{link}" target="_blank">CONSULTAR POR WHATSAPP</a>
        </div>
      </article>
    """

    cards += card_html

    if p["descuento"] > 0:
        cards_ofertas += card_html

if not cards:
    cards = """
      <p class="sin-productos">
        No se encontraron productos visibles. Revisá el archivo productos.xlsx y la columna Stock.
      </p>
    """


if not cards_ofertas:
    cards_ofertas = """
      <p class="sin-productos">
        No hay productos en oferta. Para agregarlos, completá la columna Oferta del Excel con textos como "oferta 10%" u "oferta 15%".
      </p>
    """

html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Mucho de Ti | Lencería & Corsetería</title>
  <link rel="stylesheet" href="style.css">
</head>
<body>

  <section class="topbar">
    <span>🚚 Envíos a todo el país</span>
    <span>💳 3 cuotas sin interés</span>
    <span>💬 Atención personalizada por WhatsApp</span>
  </section>

  <header class="navbar">
    <a href="#inicio" class="brand-text">Mucho de Ti</a>
    <button class="menu-toggle" id="menu-toggle" aria-label="Abrir menú" type="button">☰</button>

    <nav class="menu">
      <a href="#inicio">INICIO</a>
      <div class="menu-dropdown">
        <a href="#productos" id="menu-productos">PRODUCTOS</a>
        <div class="dropdown-content">
          {nav_dropdown}
        </div>
      </div>
      <a href="#ofertas" id="menu-ofertas">OFERTAS</a>
      <a href="#sobre-nosotros" id="menu-sobre-nosotros">SOBRE NOSOTROS</a>
      <a href="#contacto">CONTACTO</a>
    </nav>
  </header>

  <main id="inicio">
    <section class="hero">
      <div class="hero-content">
        <h1 class="hero-logo-text">MUCHO<br>DE TI</h1>
        <div class="hero-line"></div>
        <div class="hero-subtitle">LENCERÍA &amp; CORSETERÍA</div>
        <p>
          Diseños que realzan tu belleza y te hacen sentir única.
          Calidad, comodidad y estilo en cada detalle.
        </p>
        <a class="btn-principal" href="#productos">VER COLECCIÓN</a>
      </div>
    </section>

    <section class="beneficios">
      <div class="beneficio">
        <span class="beneficio-icono">▧</span>
        <div>
          <strong>ENVÍOS A TODO EL PAÍS</strong>
          <small>Rápidos y seguros</small>
        </div>
      </div>

      <div class="beneficio">
        <span class="beneficio-icono">♡</span>
        <div>
          <strong>COMPRA SEGURA</strong>
          <small>Tus datos protegidos</small>
        </div>
      </div>

      <div class="beneficio">
        <span class="beneficio-icono">▭</span>
        <div>
          <strong>3 CUOTAS SIN INTERÉS</strong>
          <small>Con tarjetas seleccionadas</small>
        </div>
      </div>

      <div class="beneficio">
        <span class="beneficio-icono">☏</span>
        <div>
          <strong>ATENCIÓN PERSONALIZADA</strong>
          <small>Te ayudamos por WhatsApp</small>
        </div>
      </div>
    </section>

    <section class="instagram-box">
      <div class="instagram-contenido">
        <small>SEGUINOS EN INSTAGRAM</small>
        <h2>@muchodetilenceria</h2>
        <p>
          Conocé novedades, ingresos, colecciones y propuestas seleccionadas especialmente para vos.
        </p>
        <a class="btn-instagram" href="https://www.instagram.com/muchodetilenceria/" target="_blank">
          VER INSTAGRAM
        </a>
      </div>
    </section>

    <section class="sobre-nosotros" id="sobre-nosotros" style="display:none;">
      <div class="sobre-contenedor">
        <div class="sobre-texto">
          <small>NUESTRA ESENCIA</small>
          <h2>Mucho de Ti</h2>
          <h3>Lencería pensada para acompañarte, realzarte y hacerte sentir única.</h3>

          <p>
            En Mucho de Ti creemos que la lencería no es solo una prenda: es una forma de conectar
            con vos misma, de sentirte cómoda, segura y auténtica en cada momento.
          </p>

          <p>
            Seleccionamos cuidadosamente lencería, corsetería y prendas íntimas de marcas reconocidas,
            priorizando la calidad, el calce, la comodidad y los detalles que hacen la diferencia.
            Cada colección está pensada para acompañar distintos cuerpos, estilos y necesidades.
          </p>

          <p>
            Nuestro objetivo es brindarte una experiencia cercana y personalizada. Te ayudamos a elegir
            el modelo, talle o diseño ideal para que encuentres esa prenda que no solo te guste,
            sino que también te haga sentir bien.
          </p>

          <a class="btn-principal" href="#productos" id="btn-sobre-productos">VER PRODUCTOS</a>
        </div>

        <div class="sobre-destacados">
          <div class="sobre-card">
            <span>♡</span>
            <h4>Selección cuidada</h4>
            <p>Trabajamos con productos elegidos por su diseño, comodidad y calidad.</p>
          </div>

          <div class="sobre-card">
            <span>✦</span>
            <h4>Atención personalizada</h4>
            <p>Te acompañamos para encontrar la prenda, el talle y el estilo ideal para vos.</p>
          </div>

          <div class="sobre-card">
            <span>▧</span>
            <h4>Marcas reconocidas</h4>
            <p>Contamos con variedad de marcas, modelos y propuestas para diferentes gustos.</p>
          </div>
        </div>
      </div>
    </section>

    <section class="brand-nav" id="marcas">
      {nav_marcas}
    </section>


    <section class="ofertas" id="ofertas">
      <small>OPORTUNIDADES ESPECIALES</small>
      <h2>OFERTAS</h2>
      <div class="grid grid-ofertas" id="grid-ofertas">
        {cards_ofertas}
      </div>
    </section>

    <section class="productos" id="productos">
      <small id="subtitulo-productos">SELECCIÓN ESPECIAL</small>
      <h2 id="titulo-productos">PRODUCTOS DESTACADOS</h2>

      <div class="herramientas-productos">
        <input id="buscador" class="buscador-productos" type="search" placeholder="Buscar por nombre, código o descripción...">
        <p id="contador-resultados" class="contador-resultados">Mostrando productos</p>
      </div>

      <div class="grid" id="grid-productos">
        {cards}
      </div>
      <p class="sin-productos" id="mensaje-sin-productos" style="display:none;">
        No hay productos para mostrar en esta selección.
      </p>
    </section>
  </main>

  <section class="contacto" id="contacto">
    <div class="contacto-contenedor">
      <div class="contacto-info">
        <small>CONTACTO</small>
        <h2>Estamos para ayudarte</h2>

        <p>
          Si tenés dudas sobre talles, modelos, disponibilidad o querés recibir atención personalizada,
          podés comunicarte con nosotras o visitarnos en nuestro local.
        </p>

        <div class="contacto-datos">
          <div class="contacto-item">
            <strong>Dirección</strong>
            <span>Uruguay 473, C1015ABI<br>Ciudad Autónoma de Buenos Aires, Argentina</span>
          </div>

          <div class="contacto-item">
            <strong>Teléfono</strong>
            <span>011 4372-2199</span>
          </div>

          <div class="contacto-item">
            <strong>Email</strong>
            <span>
              <a href="mailto:muchodetilenceria@gmail.com">
                muchodetilenceria@gmail.com
              </a>
            </span>
          </div>

          <div class="contacto-item">
            <strong>Celular / WhatsApp</strong>
            <span>+54 9 11 6350-9142</span>
          </div>

          <div class="contacto-item">
            <strong>Horario</strong>
            <span>Lunes a Viernes: 09:00 - 18:30</span>
          </div>
        </div>

        <div class="contacto-botones">
          <a class="btn-principal" href="https://wa.me/{WHATSAPP_NUMERO}" target="_blank">
            ESCRIBIR POR WHATSAPP
          </a>

          <a class="btn-secundario" href="https://www.google.com/maps/search/?api=1&query=Uruguay+473+C1015ABI+Ciudad+Autonoma+de+Buenos+Aires+Argentina" target="_blank">
            VER EN GOOGLE MAPS
          </a>
        </div>
      </div>

      <div class="contacto-mapa">
        <iframe
          src="https://www.google.com/maps?q=Uruguay%20473%2C%20C1015ABI%20Ciudad%20Aut%C3%B3noma%20de%20Buenos%20Aires%2C%20Argentina&output=embed"
          width="100%"
          height="420"
          style="border:0;"
          allowfullscreen=""
          loading="lazy"
          referrerpolicy="no-referrer-when-downgrade">
        </iframe>
      </div>
    </div>
  </section>

  <button class="carrito-flotante" id="carrito-flotante" type="button" aria-label="Abrir carrito">
    🛍️
    <span id="carrito-cantidad" class="carrito-cantidad">0</span>
  </button>

  <aside class="carrito-panel" id="carrito-panel" aria-hidden="true">
    <div class="carrito-header">
      <div>
        <small>TU PEDIDO</small>
        <h2>Carrito</h2>
      </div>
      <button class="carrito-cerrar" id="carrito-cerrar" type="button" aria-label="Cerrar carrito">×</button>
    </div>

    <div class="carrito-items" id="carrito-items">
      <p class="carrito-vacio">Todavía no agregaste productos.</p>
    </div>

    <div class="carrito-total">
      <span>Total estimado</span>
      <strong id="carrito-total">$0</strong>
    </div>

    <div class="carrito-datos">
      <input id="cliente-nombre" type="text" placeholder="Tu nombre">
      <input id="cliente-entrega" type="text" placeholder="Retiro / envío / zona">
      <textarea id="cliente-notas" rows="3" placeholder="Observaciones, talle, color o consultas"></textarea>
    </div>

    <button class="finalizar-whatsapp" id="finalizar-whatsapp" type="button">
      FINALIZAR POR WHATSAPP
    </button>
  </aside>

  <div class="carrito-overlay" id="carrito-overlay"></div>

  <a class="wa-flotante" href="https://wa.me/{WHATSAPP_NUMERO}" target="_blank" aria-label="Escribir por WhatsApp">💬</a>
  <button class="volver-arriba" id="volver-arriba" type="button" aria-label="Volver arriba">⌃</button>

  <div class="lightbox" id="lightbox" aria-hidden="true">
    <button class="lightbox-cerrar" id="lightbox-cerrar" type="button" aria-label="Cerrar imagen">×</button>
    <img id="lightbox-img" src="" alt="Producto ampliado">
  </div>

  <script>
    const menuToggle = document.getElementById('menu-toggle');
    const menu = document.querySelector('.menu');
    const menuDropdown = document.querySelector('.menu-dropdown');

    const filtrosMarca = document.querySelectorAll('.filtro-marca');
    const filtrosCategoria = document.querySelectorAll('.filtro-categoria');
    const filtrosDestacados = document.querySelectorAll('.dropdown-destacados');
    const filtrosOfertas = document.querySelectorAll('.dropdown-ofertas');

    const cardsProductos = document.querySelectorAll('#grid-productos .card');
    const tituloProductos = document.getElementById('titulo-productos');
    const subtituloProductos = document.getElementById('subtitulo-productos');
    const menuProductos = document.getElementById('menu-productos');
    const menuOfertas = document.getElementById('menu-ofertas');
    const menuSobreNosotros = document.getElementById('menu-sobre-nosotros');
    const seccionSobreNosotros = document.getElementById('sobre-nosotros');
    const seccionProductos = document.getElementById('productos');
    const mensajeSinProductos = document.getElementById('mensaje-sin-productos');
    const btnSobreProductos = document.getElementById('btn-sobre-productos');
    const buscador = document.getElementById('buscador');
    const contadorResultados = document.getElementById('contador-resultados');

    const botonVolverArriba = document.getElementById('volver-arriba');
    const lightbox = document.getElementById('lightbox');
    const lightboxImg = document.getElementById('lightbox-img');
    const lightboxCerrar = document.getElementById('lightbox-cerrar');

    let filtroActual = {{
      tipo: 'destacados',
      titulo: 'PRODUCTOS DESTACADOS',
      subtitulo: 'SELECCIÓN ESPECIAL',
      descripcion: 'productos destacados'
    }};

    if (menuToggle && menu) {{
      menuToggle.addEventListener('click', function() {{
        menu.classList.toggle('abierto');
      }});
    }}

    function cerrarMenuMobile() {{
      if (menu) {{
        menu.classList.remove('abierto');
      }}

      if (menuDropdown) {{
        menuDropdown.classList.remove('abierto');
      }}
    }}

    function limpiarActivos() {{
      filtrosMarca.forEach(boton => boton.classList.remove('active'));
      filtrosCategoria.forEach(boton => boton.classList.remove('active'));
      filtrosDestacados.forEach(boton => boton.classList.remove('active'));
      filtrosOfertas.forEach(boton => boton.classList.remove('active'));
    }}

    function ocultarSobreNosotros() {{
      if (seccionSobreNosotros) {{
        seccionSobreNosotros.style.display = 'none';
      }}
      if (seccionProductos) {{
        seccionProductos.style.display = 'block';
      }}
    }}

    function cardCumpleFiltro(card) {{
      if (card.dataset.stock !== 'si') {{
        return false;
      }}
      if (filtroActual.tipo === 'destacados') {{
        return card.dataset.destacado === 'si';
      }}
      if (filtroActual.tipo === 'marca') {{
        return card.dataset.marca === filtroActual.marca;
      }}
      if (filtroActual.tipo === 'categoria') {{
        return card.dataset.grupo === filtroActual.grupo && card.dataset.categoria === filtroActual.categoria;
      }}
      return true;
    }}

    function aplicarFiltrosProductos() {{
      const q = buscador ? buscador.value.trim().toLowerCase() : '';
      let visibles = 0;
      cardsProductos.forEach(card => {{
        const texto = card.textContent.toLowerCase();
        const coincideBusqueda = q === '' || texto.includes(q);
        const visible = cardCumpleFiltro(card) && coincideBusqueda;
        card.style.display = visible ? 'block' : 'none';
        if (visible) visibles++;
      }});
      if (mensajeSinProductos) {{
        mensajeSinProductos.style.display = visibles === 0 ? 'block' : 'none';
      }}
      if (contadorResultados) {{
        const palabra = visibles === 1 ? 'producto' : 'productos';
        const extraBusqueda = q ? ` con “${{q}}”` : '';
        contadorResultados.textContent = `Mostrando ${{visibles}} ${{palabra}} de ${{filtroActual.descripcion}}${{extraBusqueda}}`;
      }}
    }}

    function mostrarSobreNosotros() {{
      if (seccionProductos) seccionProductos.style.display = 'none';
      if (seccionSobreNosotros) {{
        seccionSobreNosotros.style.display = 'block';
        seccionSobreNosotros.scrollIntoView({{ behavior: 'smooth' }});
      }}
      limpiarActivos();
      cerrarMenuMobile();
    }}

    function mostrarDestacados() {{
      ocultarSobreNosotros();
      filtroActual = {{
        tipo: 'destacados',
        titulo: 'PRODUCTOS DESTACADOS',
        subtitulo: 'SELECCIÓN ESPECIAL',
        descripcion: 'productos destacados'
      }};
      limpiarActivos();
      subtituloProductos.textContent = filtroActual.subtitulo;
      tituloProductos.textContent = filtroActual.titulo;
      aplicarFiltrosProductos();
    }}

    function mostrarMarca(marca) {{
      ocultarSobreNosotros();
      filtroActual = {{
        tipo: 'marca',
        marca: marca,
        titulo: marca.toUpperCase(),
        subtitulo: 'PRODUCTOS DE LA MARCA',
        descripcion: `productos de ${{marca}}`
      }};
      subtituloProductos.textContent = filtroActual.subtitulo;
      tituloProductos.textContent = filtroActual.titulo;
      aplicarFiltrosProductos();
    }}

    function mostrarCategoria(grupo, categoria) {{
      ocultarSobreNosotros();
      filtroActual = {{
        tipo: 'categoria',
        grupo: grupo,
        categoria: categoria,
        titulo: categoria.toUpperCase(),
        subtitulo: grupo.toUpperCase(),
        descripcion: `${{categoria}}`
      }};
      subtituloProductos.textContent = filtroActual.subtitulo;
      tituloProductos.textContent = filtroActual.titulo;
      aplicarFiltrosProductos();
    }}

    function scrollAProductos() {{
      const productos = document.getElementById('productos');
      if (!productos) return;

      const offset = window.innerWidth <= 820 ? 220 : 130;
      const posicion = productos.getBoundingClientRect().top + window.pageYOffset - offset;

      window.scrollTo({{
        top: posicion,
        behavior: 'smooth'
      }});
    }}

    if (buscador) buscador.addEventListener('input', aplicarFiltrosProductos);

    filtrosDestacados.forEach(boton => {{
      boton.addEventListener('click', function(e) {{
        e.preventDefault();
        limpiarActivos();
        this.classList.add('active');
    
    /* ===== CARRITO POR WHATSAPP ===== */
    const carritoKey = 'muchoDeTiCarrito';
    const carritoPanel = document.getElementById('carrito-panel');
    const carritoOverlay = document.getElementById('carrito-overlay');
    const carritoAbrir = document.getElementById('carrito-flotante');
    const carritoCerrar = document.getElementById('carrito-cerrar');
    const carritoItems = document.getElementById('carrito-items');
    const carritoCantidad = document.getElementById('carrito-cantidad');
    const carritoTotal = document.getElementById('carrito-total');
    const finalizarWhatsapp = document.getElementById('finalizar-whatsapp');

    let carritoMemoria = [];

    function leerCarrito() {{
      try {{
        const guardado = localStorage.getItem(carritoKey);
        carritoMemoria = guardado ? JSON.parse(guardado) : carritoMemoria;
        return carritoMemoria || [];
      }} catch (e) {{
        return carritoMemoria || [];
      }}
    }}

    function guardarCarrito(carrito) {{
      carritoMemoria = carrito;
      try {{
        localStorage.setItem(carritoKey, JSON.stringify(carrito));
      }} catch (e) {{
        // Si el navegador bloquea localStorage, el carrito sigue funcionando en memoria.
      }}
      renderizarCarrito();
    }}

    function precioNumero(precio) {{
      let texto = String(precio || '').replace('$', '').replace(/\\s/g, '');

      if (!texto) return 0;

      if (texto.includes(',')) {{
        texto = texto.replace(/\\./g, '').replace(',', '.');
      }} else {{
        texto = texto.replace(/\\./g, '');
      }}

      const valor = parseFloat(texto);
      return isNaN(valor) ? 0 : valor;
    }}

    function formatoPesos(valor) {{
      return '$' + Math.round(valor).toLocaleString('es-AR');
    }}

    function abrirCarrito() {{
      if (!carritoPanel || !carritoOverlay) return;
      carritoPanel.classList.add('abierto');
      carritoOverlay.classList.add('abierto');
      carritoPanel.setAttribute('aria-hidden', 'false');
    }}

    function cerrarCarrito() {{
      if (!carritoPanel || !carritoOverlay) return;
      carritoPanel.classList.remove('abierto');
      carritoOverlay.classList.remove('abierto');
      carritoPanel.setAttribute('aria-hidden', 'true');
    }}

    function agregarAlCarrito(producto) {{
      const carrito = leerCarrito();
      const existente = carrito.find(item => item.codigo === producto.codigo);

      if (existente) {{
        existente.cantidad += 1;
      }} else {{
        carrito.push({{ ...producto, cantidad: 1 }});
      }}

      guardarCarrito(carrito);
      abrirCarrito();
    }}

    function cambiarCantidad(codigo, delta) {{
      const carrito = leerCarrito();
      const item = carrito.find(producto => producto.codigo === codigo);

      if (!item) return;

      item.cantidad += delta;

      const actualizado = carrito.filter(producto => producto.cantidad > 0);
      guardarCarrito(actualizado);
    }}

    function quitarDelCarrito(codigo) {{
      const carrito = leerCarrito().filter(producto => producto.codigo !== codigo);
      guardarCarrito(carrito);
    }}

    function renderizarCarrito() {{
      const carrito = leerCarrito();
      const cantidadTotal = carrito.reduce((total, item) => total + item.cantidad, 0);
      const total = carrito.reduce((sum, item) => sum + precioNumero(item.precio) * item.cantidad, 0);

      if (carritoCantidad) {{
        carritoCantidad.textContent = cantidadTotal;
      }}

      if (carritoTotal) {{
        carritoTotal.textContent = formatoPesos(total);
      }}

      if (!carritoItems) return;

      if (carrito.length === 0) {{
        carritoItems.innerHTML = '<p class="carrito-vacio">Todavía no agregaste productos.</p>';
        return;
      }}

      carritoItems.innerHTML = carrito.map(item => `
        <div class="carrito-item">
          <img src="${{item.imagen}}" alt="${{item.nombre}}">
          <div class="carrito-item-info">
            <strong>${{item.nombre}}</strong>
            <span>${{item.marca}} · Cód. ${{item.codigo}}</span>
            <small>${{item.precio}}</small>
            <div class="carrito-controles">
              <button type="button" data-accion="menos" data-codigo="${{item.codigo}}">−</button>
              <span>${{item.cantidad}}</span>
              <button type="button" data-accion="mas" data-codigo="${{item.codigo}}">+</button>
              <button type="button" class="quitar" data-accion="quitar" data-codigo="${{item.codigo}}">Quitar</button>
            </div>
          </div>
        </div>
      `).join('');
    }}

    document.querySelectorAll('.btn-carrito-card').forEach(boton => {{
      boton.addEventListener('click', function(e) {{
        e.preventDefault();
        e.stopPropagation();
        agregarAlCarrito({{
          codigo: this.dataset.codigo,
          nombre: this.dataset.nombre,
          marca: this.dataset.marca,
          precio: this.dataset.precio,
          imagen: this.dataset.imagen
        }});
      }});
    }});

    document.addEventListener('click', function(e) {{
      const botonCarrito = e.target.closest('.btn-carrito-card');
      if (!botonCarrito) return;

      e.preventDefault();
      e.stopPropagation();

      agregarAlCarrito({{
        codigo: botonCarrito.dataset.codigo,
        nombre: botonCarrito.dataset.nombre,
        marca: botonCarrito.dataset.marca,
        precio: botonCarrito.dataset.precio,
        imagen: botonCarrito.dataset.imagen
      }});
    }});

    if (carritoAbrir) {{
      carritoAbrir.addEventListener('click', abrirCarrito);
    }}

    if (carritoCerrar) {{
      carritoCerrar.addEventListener('click', cerrarCarrito);
    }}

    if (carritoOverlay) {{
      carritoOverlay.addEventListener('click', cerrarCarrito);
    }}

    if (carritoItems) {{
      carritoItems.addEventListener('click', function(e) {{
        const boton = e.target.closest('button');
        if (!boton) return;

        const codigo = boton.dataset.codigo;
        const accion = boton.dataset.accion;

        if (accion === 'mas') cambiarCantidad(codigo, 1);
        if (accion === 'menos') cambiarCantidad(codigo, -1);
        if (accion === 'quitar') quitarDelCarrito(codigo);
      }});
    }}

    if (finalizarWhatsapp) {{
      finalizarWhatsapp.addEventListener('click', function() {{
        const carrito = leerCarrito();

        if (carrito.length === 0) {{
          alert('Agregá al menos un producto al carrito.');
          return;
        }}

        const nombre = document.getElementById('cliente-nombre')?.value.trim() || '';
        const entrega = document.getElementById('cliente-entrega')?.value.trim() || '';
        const notas = document.getElementById('cliente-notas')?.value.trim() || '';
        const total = carrito.reduce((sum, item) => sum + precioNumero(item.precio) * item.cantidad, 0);

        const lineas = [
          'Hola, quiero hacer este pedido:',
          '',
          ...carrito.map(item =>
            `${{item.cantidad}} x ${{item.nombre}}\\nMarca: ${{item.marca}}\\nCódigo: ${{item.codigo}}\\nPrecio: ${{item.precio}}`
          ),
          '',
          `Total estimado: ${{formatoPesos(total)}}`,
          '',
          'Mis datos:',
          `Nombre: ${{nombre}}`,
          `Entrega/retiro: ${{entrega}}`,
          `Observaciones: ${{notas}}`
        ];

        const mensaje = encodeURIComponent(lineas.join("\\\\n\\\\n"));
        window.open(`https://wa.me/{WHATSAPP_NUMERO}?text=${{mensaje}}`, '_blank');
      }});
    }}

    renderizarCarrito();


    mostrarDestacados();
        cerrarMenuMobile();
        setTimeout(scrollAProductos, 120);
      }});
    }});

    filtrosOfertas.forEach(boton => {{
      boton.addEventListener('click', function(e) {{
        e.preventDefault();
        limpiarActivos();
        this.classList.add('active');
        document.getElementById('ofertas').scrollIntoView({{ behavior: 'smooth' }});
        cerrarMenuMobile();
      }});
    }});

    filtrosMarca.forEach(boton => {{
      boton.addEventListener('click', function(e) {{
        e.preventDefault();
        limpiarActivos();
        this.classList.add('active');
        mostrarMarca(this.dataset.marca);
        cerrarMenuMobile();
        setTimeout(scrollAProductos, 120);
      }});
    }});

    filtrosCategoria.forEach(boton => {{
      boton.addEventListener('click', function(e) {{
        e.preventDefault();
        limpiarActivos();
        this.classList.add('active');
        mostrarCategoria(this.dataset.grupo, this.dataset.categoria);
        cerrarMenuMobile();
        setTimeout(scrollAProductos, 120);
      }});
    }});

    if (menuProductos) {{
      menuProductos.addEventListener('click', function(e) {{
        if (window.innerWidth <= 820) {{
          e.preventDefault();

          if (menuDropdown) {{
            menuDropdown.classList.toggle('abierto');
          }}

          return;
        }}

        e.preventDefault();
        mostrarDestacados();
        cerrarMenuMobile();
        setTimeout(scrollAProductos, 120);
      }});
    }}

    if (menuOfertas) {{
      menuOfertas.addEventListener('click', function(e) {{
        e.preventDefault();
        limpiarActivos();
        document.getElementById('ofertas').scrollIntoView({{ behavior: 'smooth' }});
        cerrarMenuMobile();
      }});
    }}

    if (menuSobreNosotros) {{
      menuSobreNosotros.addEventListener('click', function(e) {{
        e.preventDefault();
        mostrarSobreNosotros();
      }});
    }}

    if (btnSobreProductos) {{
      btnSobreProductos.addEventListener('click', function(e) {{
        e.preventDefault();
        mostrarDestacados();
        cerrarMenuMobile();
        setTimeout(scrollAProductos, 120);
      }});
    }}

    if (botonVolverArriba) {{
      window.addEventListener('scroll', function() {{
        if (window.scrollY > 400) botonVolverArriba.classList.add('visible');
        else botonVolverArriba.classList.remove('visible');
      }});
      botonVolverArriba.addEventListener('click', function() {{
        window.scrollTo({{ top: 0, behavior: 'smooth' }});
      }});
    }}

    document.querySelectorAll('.producto-img').forEach(img => {{
      img.addEventListener('click', function() {{
        if (!lightbox || !lightboxImg) return;
        lightboxImg.src = this.src;
        lightboxImg.alt = this.alt || 'Producto ampliado';
        lightbox.classList.add('abierto');
        lightbox.setAttribute('aria-hidden', 'false');
      }});
    }});

    function cerrarLightbox() {{
      if (!lightbox || !lightboxImg) return;
      lightbox.classList.remove('abierto');
      lightbox.setAttribute('aria-hidden', 'true');
      lightboxImg.src = '';
    }}

    if (lightboxCerrar) lightboxCerrar.addEventListener('click', cerrarLightbox);
    if (lightbox) {{
      lightbox.addEventListener('click', function(e) {{
        if (e.target === lightbox) cerrarLightbox();
      }});
    }}
    document.addEventListener('keydown', function(e) {{
      if (e.key === 'Escape') cerrarLightbox();
    }});

    mostrarDestacados();
  </script>


  <script>
    /* CARRITO FALLBACK INDEPENDIENTE */
    (function() {{
      const WHATSAPP = "{WHATSAPP_NUMERO}";
      let carritoFallback = [];

      function leer() {{
        try {{
          carritoFallback = JSON.parse(localStorage.getItem("muchoDeTiCarrito")) || carritoFallback || [];
        }} catch (e) {{
          carritoFallback = carritoFallback || [];
        }}
        return carritoFallback;
      }}

      function guardar(carrito) {{
        carritoFallback = carrito;
        try {{
          localStorage.setItem("muchoDeTiCarrito", JSON.stringify(carrito));
        }} catch (e) {{}}
        render();
      }}

      function precioNumero(precio) {{
        let texto = String(precio || "").replace("$", "").replace(/\\s/g, "");
        if (!texto) return 0;
        if (texto.includes(",")) {{
          texto = texto.replace(/\\./g, "").replace(",", ".");
        }} else {{
          texto = texto.replace(/\\./g, "");
        }}
        const valor = parseFloat(texto);
        return isNaN(valor) ? 0 : valor;
      }}

      function formatoPesos(valor) {{
        return "$" + Math.round(valor).toLocaleString("es-AR");
      }}

      function abrir() {{
        const panel = document.getElementById("carrito-panel");
        const overlay = document.getElementById("carrito-overlay");
        if (panel) {{
          panel.classList.add("abierto");
          panel.setAttribute("aria-hidden", "false");
        }}
        if (overlay) overlay.classList.add("abierto");
      }}

      function cerrar() {{
        const panel = document.getElementById("carrito-panel");
        const overlay = document.getElementById("carrito-overlay");
        if (panel) {{
          panel.classList.remove("abierto");
          panel.setAttribute("aria-hidden", "true");
        }}
        if (overlay) overlay.classList.remove("abierto");
      }}

      function render() {{
        const carrito = leer();
        const cantidad = carrito.reduce((total, item) => total + Number(item.cantidad || 0), 0);
        const total = carrito.reduce((sum, item) => sum + precioNumero(item.precio) * Number(item.cantidad || 0), 0);

        const cantidadEl = document.getElementById("carrito-cantidad");
        const totalEl = document.getElementById("carrito-total");
        const itemsEl = document.getElementById("carrito-items");

        if (cantidadEl) cantidadEl.textContent = cantidad;
        if (totalEl) totalEl.textContent = formatoPesos(total);

        if (!itemsEl) return;

        if (!carrito.length) {{
          itemsEl.innerHTML = '<p class="carrito-vacio">Todavía no agregaste productos.</p>';
          return;
        }}

        itemsEl.innerHTML = carrito.map(item => `
          <div class="carrito-item">
            <img src="${{item.imagen}}" alt="${{item.nombre}}">
            <div class="carrito-item-info">
              <strong>${{item.nombre}}</strong>
              <span>${{item.marca}} · Cód. ${{item.codigo}}</span>
              <small>${{item.precio}}</small>
              <div class="carrito-controles">
                <button type="button" data-cart-action="menos" data-codigo="${{item.codigo}}">−</button>
                <span>${{item.cantidad}}</span>
                <button type="button" data-cart-action="mas" data-codigo="${{item.codigo}}">+</button>
                <button type="button" class="quitar" data-cart-action="quitar" data-codigo="${{item.codigo}}">Quitar</button>
              </div>
            </div>
          </div>
        `).join("");
      }}

      window.agregarAlCarritoDesdeBoton = function(boton) {{
        if (!boton) return;

        const producto = {{
          codigo: boton.dataset.codigo || "",
          nombre: boton.dataset.nombre || "",
          marca: boton.dataset.marca || "",
          precio: boton.dataset.precio || "",
          imagen: boton.dataset.imagen || ""
        }};

        const carrito = leer();
        const existente = carrito.find(item => item.codigo === producto.codigo);

        if (existente) {{
          existente.cantidad += 1;
        }} else {{
          carrito.push({{ ...producto, cantidad: 1 }});
        }}

        guardar(carrito);
        abrir();
      }};

      document.addEventListener("click", function(e) {{
        const botonAgregar = e.target.closest(".btn-carrito-card");
        if (botonAgregar) {{
          e.preventDefault();
          e.stopPropagation();
          window.agregarAlCarritoDesdeBoton(botonAgregar);
          return;
        }}

        const abrirBtn = e.target.closest("#carrito-flotante");
        if (abrirBtn) {{
          e.preventDefault();
          abrir();
          return;
        }}

        const cerrarBtn = e.target.closest("#carrito-cerrar, #carrito-overlay");
        if (cerrarBtn) {{
          e.preventDefault();
          cerrar();
          return;
        }}

        const control = e.target.closest("[data-cart-action]");
        if (control) {{
          const codigo = control.dataset.codigo;
          const accion = control.dataset.cartAction;
          let carrito = leer();

          if (accion === "mas") {{
            const item = carrito.find(p => p.codigo === codigo);
            if (item) item.cantidad += 1;
          }}

          if (accion === "menos") {{
            const item = carrito.find(p => p.codigo === codigo);
            if (item) item.cantidad -= 1;
            carrito = carrito.filter(p => p.cantidad > 0);
          }}

          if (accion === "quitar") {{
            carrito = carrito.filter(p => p.codigo !== codigo);
          }}

          guardar(carrito);
        }}
      }});

      const finalizar = document.getElementById("finalizar-whatsapp");
      if (finalizar) {{
        finalizar.addEventListener("click", function() {{
          const carrito = leer();

          if (!carrito.length) {{
            alert("Agregá al menos un producto al carrito.");
            return;
          }}

          const nombre = document.getElementById("cliente-nombre")?.value.trim() || "";
          const entrega = document.getElementById("cliente-entrega")?.value.trim() || "";
          const notas = document.getElementById("cliente-notas")?.value.trim() || "";
          const total = carrito.reduce((sum, item) => sum + precioNumero(item.precio) * Number(item.cantidad || 0), 0);

          const lineas = [
            "Hola, quiero hacer este pedido:",
            "",
            ...carrito.map(item =>
              `${{item.cantidad}} x ${{item.nombre}}\\nMarca: ${{item.marca}}\\nCódigo: ${{item.codigo}}\\nPrecio: ${{item.precio}}`
            ),
            "",
            `Total estimado: ${{formatoPesos(total)}}`,
            "",
            "Mis datos:",
            `Nombre: ${{nombre}}`,
            `Entrega/retiro: ${{entrega}}`,
            `Observaciones: ${{notas}}`
          ];

          window.open(`https://wa.me/${{WHATSAPP}}?text=${{encodeURIComponent(lineas.join("\\n\\n"))}}`, "_blank");
        }});
      }}

      render();
    }})();
  </script>

</body>
</html>
"""

with open("index.html", "w", encoding="utf-8") as f:
    f.write(html)

print("Página generada desde Excel: index.html")
print(f"Productos visibles según stock: {len(productos)}")
print(f"Productos destacados visibles: {len(productos_destacados)}")
print(f"Productos en oferta: {len([p for p in productos if p['descuento'] > 0])}")
print(f"Cards generadas en HTML: {len(productos)}")
